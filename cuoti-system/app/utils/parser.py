# -*- coding: utf-8 -*-
"""文档解析器 - 支持Word/PDF/Excel"""
import re
import os
import json
from datetime import datetime


def parse_document(file_path):
    """根据文件类型自动选择解析器"""
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.docx':
        return parse_word_document(file_path)
    elif ext == '.pdf':
        return parse_pdf_document(file_path)
    elif ext in ['.xlsx', '.xls']:
        return parse_excel_document(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


def parse_word_document(file_path):
    """解析Word文档"""
    from docx import Document
    
    doc = Document(file_path)
    full_text = []
    
    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            full_text.append(text)
    
    return parse_text_content('\n'.join(full_text), os.path.basename(file_path))


def parse_pdf_document(file_path):
    """解析PDF文档"""
    try:
        import pdfplumber
        
        full_text = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text.append(text)
        
        return parse_text_content('\n'.join(full_text), os.path.basename(file_path))
    except Exception as e:
        return {
            'success': False,
            'error': f"PDF解析失败: {str(e)}",
            'questions': []
        }


def parse_excel_document(file_path):
    """解析Excel文档 - 固定格式"""
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    questions = []
    headers = [cell.value for cell in ws[1]]
    
    # 尝试识别列
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            h_lower = str(h).lower()
            if '题干' in h or 'stem' in h_lower or '题目' in h:
                col_map['stem'] = i
            elif '选项a' in h.lower() or 'a' == h_lower:
                col_map['option_a'] = i
            elif '选项b' in h.lower() or 'b' == h_lower:
                col_map['option_b'] = i
            elif '选项c' in h.lower() or 'c' == h_lower:
                col_map['option_c'] = i
            elif '选项d' in h.lower() or 'd' == h_lower:
                col_map['option_d'] = i
            elif '答案' in h or 'answer' in h_lower:
                col_map['answer'] = i
            elif '解析' in h or 'analysis' in h_lower or '解答' in h:
                col_map['analysis'] = i
            elif '分类' in h or 'category' in h_lower or '类型' in h:
                col_map['category'] = i
            elif '知识点' in h or 'knowledge' in h_lower:
                col_map['knowledge_point'] = i
            elif '难度' in h or 'difficulty' in h_lower:
                col_map['difficulty'] = i
    
    # 解析数据行
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row or not any(row):
            continue
        
        stem = row[col_map.get('stem', 0)] if col_map.get('stem') is not None else None
        if not stem:
            continue
        
        question = {
            'number': row_idx,
            'stem': str(stem).strip(),
            'options': {
                'A': str(row[col_map.get('option_a', 1)] or '').strip() if col_map.get('option_a') is not None else '',
                'B': str(row[col_map.get('option_b', 2)] or '').strip() if col_map.get('option_b') is not None else '',
                'C': str(row[col_map.get('option_c', 3)] or '').strip() if col_map.get('option_c') is not None else '',
                'D': str(row[col_map.get('option_d', 4)] or '').strip() if col_map.get('option_d') is not None else ''
            },
            'answer': str(row[col_map.get('answer', 5)] or '').strip().upper() if col_map.get('answer') is not None else '',
            'analysis': str(row[col_map.get('analysis', 6)] or '').strip() if col_map.get('analysis') is not None else '',
            'category': str(row[col_map.get('category')] or '').strip() if col_map.get('category') is not None else '',
            'knowledge_point': str(row[col_map.get('knowledge_point')] or '').strip() if col_map.get('knowledge_point') is not None else '',
            'difficulty': str(row[col_map.get('difficulty')] or '中等').strip() if col_map.get('difficulty') is not None else '中等',
            'parse_status': 'success'
        }
        
        questions.append(question)
    
    return {
        'success': True,
        'source_file': os.path.basename(file_path),
        'parse_time': datetime.now().isoformat(),
        'total_questions': len(questions),
        'questions': questions,
        'format': 'excel'
    }


def parse_text_content(text, source_file):
    """解析纯文本内容，提取题目"""
    questions = []
    
    # 题号正则模式
    question_patterns = [
        r'^(\d+)\s*[\.、．]\s*',  # "1." "1、"
        r'^[（\(](\d+)[）\)]\s*',  # "(1)"
        r'^第(\d+)题\s*[：:．.]?\s*',  # "第1题"
    ]
    
    # 选项正则模式
    option_patterns = [
        r'^([A-D])\s*[\.、．:：]\s*(.+)$',  # "A." "A、"
        r'^[（\(]([A-D])[）\)]\s*(.+)$',  # "(A)"
    ]
    
    # 分割文本
    lines = text.split('\n')
    
    current_question = None
    current_section = ''
    question_number = 0
    
    # 章节识别
    section_patterns = [
        r'^第[一二三四五六七八九十]+部分\s*[:：]?\s*(.+)$',
        r'^[【\[](.+?)[】\]]$',
        r'^Part\s*\d+\s*[:：]?\s*(.+)$',
    ]
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # 检查是否是章节标题
        for pattern in section_patterns:
            match = re.match(pattern, line)
            if match:
                current_section = match.group(1).strip()
                break
        
        # 检查是否是新题目
        is_new_question = False
        for pattern in question_patterns:
            match = re.match(pattern, line)
            if match:
                # 保存上一题
                if current_question:
                    questions.append(current_question)
                
                question_number += 1
                stem = re.sub(pattern, '', line).strip()
                
                current_question = {
                    'number': question_number,
                    'section': current_section,
                    'stem': stem,
                    'options': {'A': '', 'B': '', 'C': '', 'D': ''},
                    'answer': '',
                    'analysis': '',
                    'parse_status': 'success'
                }
                is_new_question = True
                break
        
        if is_new_question:
            continue
        
        # 检查是否是选项
        if current_question:
            option_found = False
            for pattern in option_patterns:
                match = re.match(pattern, line)
                if match:
                    option_letter = match.group(1).upper()
                    option_content = match.group(2).strip()
                    current_question['options'][option_letter] = option_content
                    option_found = True
                    break
            
            # 检查选项是否在同一行（A.xxx B.xxx C.xxx D.xxx）
            if not option_found:
                inline_options = re.findall(r'([A-D])\s*[\.、．:：]\s*([^A-D]+?)(?=[A-D]\s*[\.、．:：]|$)', line)
                if inline_options:
                    for opt_letter, opt_content in inline_options:
                        current_question['options'][opt_letter.upper()] = opt_content.strip()
                    option_found = True
            
            # 如果不是选项，可能是题干的延续或答案
            if not option_found:
                # 检查是否是答案行
                answer_match = re.match(r'^[答案解析]*[：:]\s*([A-D]+)', line)
                if answer_match:
                    current_question['answer'] = answer_match.group(1).upper()
                elif '解析' in line or '答案' in line:
                    # 可能是解析
                    analysis_match = re.search(r'解析[：:]\s*(.+)', line)
                    if analysis_match:
                        current_question['analysis'] = analysis_match.group(1)
                elif current_question['stem'] and not any(current_question['options'].values()):
                    # 可能是题干的延续
                    current_question['stem'] += ' ' + line
    
    # 保存最后一题
    if current_question:
        questions.append(current_question)
    
    # 验证解析结果
    for q in questions:
        if not q['stem']:
            q['parse_status'] = 'error'
            q['parse_notes'] = '题干为空'
        elif not any(q['options'].values()):
            q['parse_status'] = 'warning'
            q['parse_notes'] = '未识别到选项'
    
    return {
        'success': True,
        'source_file': source_file,
        'parse_time': datetime.now().isoformat(),
        'total_questions': len(questions),
        'questions': questions,
        'format': 'text'
    }


def validate_parsed_questions(questions):
    """验证解析结果"""
    valid_count = 0
    warning_count = 0
    error_count = 0
    
    for q in questions:
        if q.get('parse_status') == 'success':
            valid_count += 1
        elif q.get('parse_status') == 'warning':
            warning_count += 1
        else:
            error_count += 1
    
    return {
        'total': len(questions),
        'valid': valid_count,
        'warning': warning_count,
        'error': error_count
    }
