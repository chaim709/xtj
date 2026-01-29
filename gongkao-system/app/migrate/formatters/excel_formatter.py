"""Excel格式化器"""
import os
from typing import Dict, Any, List
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from app.migrate.utils import MODULE_NAMES


class ExcelFormatter:
    """Excel格式处理器"""
    
    COLUMN_NAMES = {
        'id': 'ID',
        'name': '姓名',
        'username': '用户名',
        'phone': '手机号',
        'role': '角色',
        'created_at': '创建时间',
        'updated_at': '更新时间',
        'is_active': '是否激活',
        'status': '状态',
        'batch_id': '班次ID',
        'student_id': '学员ID',
        'teacher_id': '教师ID',
        'content': '内容',
        'note': '备注',
        'description': '描述',
    }
    
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(color='FFFFFF', bold=True)
    
    @classmethod
    def save_to_file(cls, data: Dict[str, Any], file_path: str) -> None:
        dir_path = os.path.dirname(file_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path)
        
        wb = Workbook()
        ws_meta = wb.active
        ws_meta.title = '导出信息'
        cls._write_meta_sheet(ws_meta, data.get('meta', {}))
        
        export_data = data.get('data', {})
        for module_name, records in export_data.items():
            if records:
                sheet_name = MODULE_NAMES.get(module_name, module_name)[:31]
                ws = wb.create_sheet(title=sheet_name)
                cls._write_data_sheet(ws, records)
        
        wb.save(file_path)
    
    @classmethod
    def _write_meta_sheet(cls, ws, meta: Dict) -> None:
        meta_items = [
            ('数据版本', meta.get('version', '')),
            ('系统', meta.get('system', '')),
            ('导出时间', meta.get('export_time', '')),
            ('导出类型', meta.get('export_type', '')),
            ('包含模块', ', '.join(meta.get('modules', []))),
            ('总记录数', meta.get('total_records', 0)),
            ('校验和', meta.get('checksum', '')),
        ]
        for row, (key, value) in enumerate(meta_items, 1):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=str(value))
            ws.cell(row=row, column=1).font = Font(bold=True)
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
    
    @classmethod
    def _write_data_sheet(cls, ws, records: List[Dict]) -> None:
        if not records:
            return
        columns = list(records[0].keys())
        for col, column_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=cls.COLUMN_NAMES.get(column_name, column_name))
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
        for row, record in enumerate(records, 2):
            for col, column_name in enumerate(columns, 1):
                value = record.get(column_name)
                if isinstance(value, (list, dict)):
                    value = str(value)
                ws.cell(row=row, column=col, value=value)
        for col in range(1, len(columns) + 1):
            column_letter = get_column_letter(col)
            max_length = max(
                len(str(ws.cell(row=1, column=col).value or '')),
                max((len(str(ws.cell(row=row, column=col).value or '')) for row in range(2, min(102, len(records) + 2))), default=0)
            )
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    @classmethod
    def load_from_file(cls, file_path: str) -> Dict[str, Any]:
        wb = load_workbook(file_path, read_only=True)
        result = {'meta': {}, 'data': {}}
        
        if '导出信息' in wb.sheetnames:
            ws_meta = wb['导出信息']
            meta_mapping = {
                '数据版本': 'version',
                '系统': 'system',
                '导出时间': 'export_time',
                '导出类型': 'export_type',
                '包含模块': 'modules',
                '总记录数': 'total_records',
                '校验和': 'checksum',
            }
            for row in ws_meta.iter_rows(min_row=1, max_col=2, values_only=True):
                if row[0] in meta_mapping:
                    key = meta_mapping[row[0]]
                    value = row[1]
                    if key == 'modules' and isinstance(value, str):
                        value = [m.strip() for m in value.split(',')]
                    elif key == 'total_records':
                        value = int(value) if value else 0
                    result['meta'][key] = value
        
        reverse_module_names = {v: k for k, v in MODULE_NAMES.items()}
        reverse_column_names = {v: k for k, v in cls.COLUMN_NAMES.items()}
        
        for sheet_name in wb.sheetnames:
            if sheet_name == '导出信息':
                continue
            module_name = reverse_module_names.get(sheet_name, sheet_name)
            ws = wb[sheet_name]
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
                header = reverse_column_names.get(cell, cell)
                headers.append(header)
            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    record = dict(zip(headers, row))
                    records.append(record)
            if records:
                result['data'][module_name] = records
        
        wb.close()
        return result
    
    @staticmethod
    def validate_file(file_path: str) -> tuple:
        if not os.path.exists(file_path):
            return False, f'文件不存在: {file_path}', None
        if not file_path.endswith(('.xlsx', '.xls')):
            return False, '文件必须是.xlsx或.xls格式', None
        try:
            wb = load_workbook(file_path, read_only=True)
            wb.close()
            return True, None, None
        except Exception as e:
            return False, f'读取文件失败: {str(e)}', None
